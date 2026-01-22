from openpyxl import Workbook

def create_tos_bridge():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Symbols in column A
    ws['A1'] = "MES"
    ws['A2'] = "MGC"
    
    # EMA9 (Custom 4) in column B
    # Note: Excel formulas start with =
    # We use the symbols from column A to make it clean
    ws['B1'] = '=RTD("tos.rtd",,"CUSTOM4","/MES:XCME")'
    ws['B2'] = '=RTD("tos.rtd",,"CUSTOM4","/MGC:XCEC")'
    
    # EMA15 (Custom 6) in column C
    ws['C1'] = '=RTD("tos.rtd",,"CUSTOM6","/MES:XCME")'
    ws['C2'] = '=RTD("tos.rtd",,"CUSTOM6","/MGC:XCEC")'
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 40
    
    save_path = "TOS_RTD_Bridge.xlsx"
    wb.save(save_path)
    print(f"Excel bridge file created successfully at: {save_path}")

if __name__ == "__main__":
    create_tos_bridge()
