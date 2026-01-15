from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
wb.remove(wb.active)

# Color scheme
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUBHEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
INPUT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
CALC_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
METRIC_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
ALERT_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')

HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
SUBHEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
INPUT_FONT = Font(name='Calibri', size=11, color='0000FF', bold=True)
CALC_FONT = Font(name='Calibri', size=11, color='000000')
METRIC_FONT = Font(name='Calibri', size=11, bold=True)

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============ DASHBOARD SHEET ============
dashboard = wb.create_sheet('Dashboard', 0)
dashboard.column_dimensions['A'].width = 25
dashboard.column_dimensions['B'].width = 18
dashboard.column_dimensions['C'].width = 18
dashboard.column_dimensions['D'].width = 18
dashboard.column_dimensions['E'].width = 18

# Title
dashboard['A1'] = 'Universal OR Strategy v5.13 - Trading Dashboard'
dashboard['A1'].font = Font(name='Calibri', size=18, bold=True, color='1F4E78')
dashboard.merge_cells('A1:E1')

dashboard['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
dashboard['A2'].font = Font(name='Calibri', size=10, italic=True)
dashboard.merge_cells('A2:E2')

# Key Metrics Section
row = 4
dashboard[f'A{row}'] = 'KEY METRICS'
dashboard[f'A{row}'].fill = HEADER_FILL
dashboard[f'A{row}'].font = HEADER_FONT
dashboard.merge_cells(f'A{row}:E{row}')

row += 1
metrics = [
    ('Total Parameters', "=COUNTA('All Parameters'!B:B)-1"),
    ('Active Strategies', '2'),
    ('Instruments Traded', '2'),
    ('Total Profit Targets', '4'),
    ('Trailing Stop Levels', '3'),
]

for metric, formula in metrics:
    dashboard[f'A{row}'] = metric
    dashboard[f'B{row}'] = formula
    dashboard[f'A{row}'].font = METRIC_FONT
    dashboard[f'B{row}'].font = CALC_FONT
    dashboard[f'B{row}'].alignment = Alignment(horizontal='right')
    row += 1

# Risk Summary
row += 1
dashboard[f'A{row}'] = 'RISK MANAGEMENT'
dashboard[f'A{row}'].fill = HEADER_FILL
dashboard[f'A{row}'].font = HEADER_FONT
dashboard.merge_cells(f'A{row}:E{row}')

row += 1
dashboard[f'A{row}'] = 'Parameter'
dashboard[f'B{row}'] = 'MES'
dashboard[f'C{row}'] = 'MGC'
for col in ['A', 'B', 'C']:
    dashboard[f'{col}{row}'].fill = SUBHEADER_FILL
    dashboard[f'{col}{row}'].font = SUBHEADER_FONT

row += 1
risk_params = [
    ('Risk Per Trade', "='All Parameters'!B11", "='All Parameters'!B11"),
    ('Min Contracts', "='All Parameters'!B13", "='All Parameters'!B15"),
    ('Max Contracts', "='All Parameters'!B14", "='All Parameters'!B16"),
    ('Min Stop (pts)', "='All Parameters'!B18", "='All Parameters'!B18"),
    ('Max Stop (pts)', "='All Parameters'!B19", "='All Parameters'!B19"),
]

for param, mes_val, mgc_val in risk_params:
    dashboard[f'A{row}'] = param
    dashboard[f'B{row}'] = mes_val
    dashboard[f'C{row}'] = mgc_val
    dashboard[f'A{row}'].font = CALC_FONT
    dashboard[f'B{row}'].font = CALC_FONT
    dashboard[f'C{row}'].font = CALC_FONT
    dashboard[f'B{row}'].number_format = '$#,##0'
    dashboard[f'C{row}'].number_format = '$#,##0'
    row += 1

# Target Distribution
row += 1
dashboard[f'A{row}'] = 'TARGET DISTRIBUTION (4-Target System)'
dashboard[f'A{row}'].fill = HEADER_FILL
dashboard[f'A{row}'].font = HEADER_FONT
dashboard.merge_cells(f'A{row}:E{row}')

row += 1
dashboard[f'A{row}'] = 'Target'
dashboard[f'B{row}'] = 'Allocation'
dashboard[f'C{row}'] = 'Type'
dashboard[f'D{row}'] = 'Distance'
for col in ['A', 'B', 'C', 'D']:
    dashboard[f'{col}{row}'].fill = SUBHEADER_FILL
    dashboard[f'{col}{row}'].font = SUBHEADER_FONT

row += 1
targets = [
    ('T1 - Quick Scalp', "='All Parameters'!B23&\"%\"", 'Fixed Points', "='All Parameters'!B20&\" pts\""),
    ('T2 - Medium', "='All Parameters'!B24&\"%\"", 'ATR Multiple', "='All Parameters'!B21&\"x ATR\""),
    ('T3 - Extended', "='All Parameters'!B25&\"%\"", 'ATR Multiple', "='All Parameters'!B22&\"x ATR\""),
    ('T4 - Runner', "='All Parameters'!B26&\"%\"", 'Trailing Stop', 'Dynamic'),
]

for target, alloc, ttype, dist in targets:
    dashboard[f'A{row}'] = target
    dashboard[f'B{row}'] = alloc
    dashboard[f'C{row}'] = ttype
    dashboard[f'D{row}'] = dist
    dashboard[f'A{row}'].font = CALC_FONT
    dashboard[f'B{row}'].font = CALC_FONT
    dashboard[f'C{row}'].font = CALC_FONT
    dashboard[f'D{row}'].font = CALC_FONT
    row += 1

# ============ ALL PARAMETERS SHEET ============
params = wb.create_sheet('All Parameters', 1)
params.column_dimensions['A'].width = 35
params.column_dimensions['B'].width = 20
params.column_dimensions['C'].width = 50

params['A1'] = 'Universal OR Strategy v5.13 - All Parameters'
params['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
params.merge_cells('A1:C1')

params['A2'] = 'Parameter Name'
params['B2'] = 'Value'
params['C2'] = 'Description'
for col in ['A', 'B', 'C']:
    params[f'{col}2'].fill = HEADER_FILL
    params[f'{col}2'].font = HEADER_FONT

param_data = [
    # Session Settings
    ('SESSION SETTINGS', '', ''),
    ('Session Start', '09:30', 'Trading session start time (OR begins here)'),
    ('Session End', '16:00', 'Trading session end time (box ends here)'),
    ('OR Timeframe (minutes)', '5', 'Duration of Opening Range window'),
    ('Time Zone', 'Eastern', 'Time zone for session times'),
    
    # Risk Management
    ('RISK MANAGEMENT', '', ''),
    ('Risk Per Trade ($)', 200, 'Maximum dollar risk per trade (when stop ≤ threshold)'),
    ('Reduced Risk ($)', 200, 'Reduced risk when stop > threshold'),
    ('MES Min Contracts', 1, 'Minimum contracts for MES'),
    ('MES Max Contracts', 30, 'Maximum contracts for MES'),
    ('MGC Min Contracts', 1, 'Minimum contracts for MGC'),
    ('MGC Max Contracts', 15, 'Maximum contracts for MGC'),
    
    # Stop Loss
    ('STOP LOSS', '', ''),
    ('Min Stop (Points)', 1.0, 'Minimum stop distance'),
    ('Max Stop (Points)', 8.0, 'Maximum stop distance'),
    ('Stop Multiplier', 0.5, 'Multiplier of OR Range for stop (0.5 = half OR)'),
    ('Stop Threshold (Points)', 5.0, 'Stop distance above which reduced risk is used'),
    
    # Profit Targets
    ('PROFIT TARGETS', '', ''),
    ('T1 Fixed Points', 1.0, 'Fixed point profit for T1 (quick scalp)'),
    ('T2 ATR Multiplier', 0.5, 'Multiplier of ATR for T2 (0.5 = half ATR)'),
    ('T3 ATR Multiplier', 1.0, 'Multiplier of ATR for T3 (1.0 = 1x ATR)'),
    ('T1 Contract %', 20, '20% for quick scalp'),
    ('T2 Contract %', 30, '30% allocation'),
    ('T3 Contract %', 30, '30% allocation'),
    ('T4 Contract % (Runner)', 20, '20% for runner/trail'),
    
    # Trailing Stops
    ('TRAILING STOPS', '', ''),
    ('BE Trigger (Points)', 2.0, 'Breakeven trigger distance'),
    ('BE Offset (Ticks)', 1, 'Breakeven offset in ticks'),
    ('Trail 1 Trigger (Points)', 3.0, 'First trailing stop trigger'),
    ('Trail 1 Distance (Points)', 2.0, 'First trailing stop distance'),
    ('Trail 2 Trigger (Points)', 4.0, 'Second trailing stop trigger'),
    ('Trail 2 Distance (Points)', 1.5, 'Second trailing stop distance'),
    ('Trail 3 Trigger (Points)', 5.0, 'Third trailing stop trigger'),
    ('Trail 3 Distance (Points)', 1.0, 'Third trailing stop distance'),
    ('Manual BE Buffer (Ticks)', 1, 'Buffer in ticks above breakeven for manual BE button'),
    
    # Display Settings
    ('DISPLAY SETTINGS', '', ''),
    ('Show Mid Line', 'TRUE', 'Show middle line in OR box'),
    ('Box Opacity (%)', 20, 'Transparency of OR box (0-100)'),
    ('Show OR Label', 'TRUE', 'Show OR high/low/range text on chart'),
    
    # RMA Settings
    ('RMA SETTINGS', '', ''),
    ('RMA Enabled', 'TRUE', 'Enable RMA (Shift+Click) entry mode'),
    ('RMA ATR Period', 14, 'ATR period for RMA calculations'),
    ('RMA Stop ATR Multiplier', 1.0, 'Multiplier of ATR for RMA stop'),
    ('RMA T1 ATR Multiplier', 0.5, 'Multiplier of ATR for RMA Target 1'),
    ('RMA T2 ATR Multiplier', 1.0, 'Multiplier of ATR for RMA Target 2'),
]

row = 3
for param_name, value, desc in param_data:
    if desc == '':  # Section header
        params[f'A{row}'] = param_name
        params.merge_cells(f'A{row}:C{row}')
        params[f'A{row}'].fill = SUBHEADER_FILL
        params[f'A{row}'].font = SUBHEADER_FONT
    else:
        params[f'A{row}'] = param_name
        params[f'B{row}'] = value
        params[f'C{row}'] = desc
        params[f'A{row}'].font = CALC_FONT
        params[f'B{row}'].font = INPUT_FONT
        params[f'C{row}'].font = CALC_FONT
        params[f'B{row}'].fill = INPUT_FILL
        params[f'B{row}'].border = thin_border
        
        if isinstance(value, (int, float)) and param_name not in ['OR Timeframe (minutes)', 'RMA ATR Period', 'BE Offset (Ticks)', 'Manual BE Buffer (Ticks)']:
            if 'Contract' in param_name or 'Opacity' in param_name or '%' in param_name:
                params[f'B{row}'].number_format = '0'
            elif '$' in param_name or 'Risk' in param_name:
                params[f'B{row}'].number_format = '$#,##0'
            else:
                params[f'B{row}'].number_format = '0.0'
    row += 1

# ============ POSITION SIZING CALCULATOR ============
calc = wb.create_sheet('Position Sizing', 2)
calc.column_dimensions['A'].width = 30
calc.column_dimensions['B'].width = 15
calc.column_dimensions['C'].width = 15
calc.column_dimensions['D'].width = 15

calc['A1'] = 'Position Sizing Calculator'
calc['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
calc.merge_cells('A1:D1')

calc['A3'] = 'INPUTS'
calc['A3'].fill = HEADER_FILL
calc['A3'].font = HEADER_FONT
calc.merge_cells('A3:D3')

row = 4
calc[f'A{row}'] = 'Instrument'
calc[f'B{row}'] = 'MES'
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = INPUT_FONT
calc[f'B{row}'].fill = INPUT_FILL

row += 1
calc[f'A{row}'] = 'Stop Distance (Points)'
calc[f'B{row}'] = 3.0
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = INPUT_FONT
calc[f'B{row}'].fill = INPUT_FILL
calc[f'B{row}'].number_format = '0.00'

row += 1
calc[f'A{row}'] = 'Point Value ($)'
calc[f'B{row}'] = 5
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = INPUT_FONT
calc[f'B{row}'].fill = INPUT_FILL
calc[f'B{row}'].number_format = '$#,##0'

row += 2
calc[f'A{row}'] = 'CALCULATIONS'
calc[f'A{row}'].fill = HEADER_FILL
calc[f'A{row}'].font = HEADER_FONT
calc.merge_cells(f'A{row}:D{row}')

row += 1
calc[f'A{row}'] = 'Risk Per Trade'
calc[f'B{row}'] = "='All Parameters'!B11"
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = CALC_FONT
calc[f'B{row}'].number_format = '$#,##0'

row += 1
calc[f'A{row}'] = 'Stop Distance ($)'
calc[f'B{row}'] = '=B5*B6'
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = CALC_FONT
calc[f'B{row}'].number_format = '$#,##0.00'

row += 1
calc[f'A{row}'] = 'Calculated Contracts'
calc[f'B{row}'] = '=FLOOR(B9/B10,1)'
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = Font(name='Calibri', size=12, bold=True, color='FF0000')
calc[f'B{row}'].number_format = '0'

row += 1
calc[f'A{row}'] = 'Min Contracts (MES)'
calc[f'B{row}'] = "='All Parameters'!B13"
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = CALC_FONT

row += 1
calc[f'A{row}'] = 'Max Contracts (MES)'
calc[f'B{row}'] = "='All Parameters'!B14"
calc[f'A{row}'].font = METRIC_FONT
calc[f'B{row}'].font = CALC_FONT

row += 1
calc[f'A{row}'] = 'Final Position Size'
calc[f'B{row}'] = '=MAX(B12,MIN(B11,B13))'
calc[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
calc[f'B{row}'].font = Font(name='Calibri', size=14, bold=True, color='0000FF')
calc[f'B{row}'].fill = ALERT_FILL
calc[f'B{row}'].number_format = '0'

row += 2
calc[f'A{row}'] = 'CONTRACT ALLOCATION'
calc[f'A{row}'].fill = HEADER_FILL
calc[f'A{row}'].font = HEADER_FONT
calc.merge_cells(f'A{row}:D{row}')

row += 1
calc[f'A{row}'] = 'Target'
calc[f'B{row}'] = 'Allocation %'
calc[f'C{row}'] = 'Contracts'
for col in ['A', 'B', 'C']:
    calc[f'{col}{row}'].fill = SUBHEADER_FILL
    calc[f'{col}{row}'].font = SUBHEADER_FONT

row += 1
calc[f'A{row}'] = 'T1 (Quick Scalp)'
calc[f'B{row}'] = "='All Parameters'!B23"
calc[f'C{row}'] = '=FLOOR(B14*B18/100,1)'
calc[f'C{row}'].number_format = '0'

row += 1
calc[f'A{row}'] = 'T2 (Medium)'
calc[f'B{row}'] = "='All Parameters'!B24"
calc[f'C{row}'] = '=FLOOR(B14*B19/100,1)'
calc[f'C{row}'].number_format = '0'

row += 1
calc[f'A{row}'] = 'T3 (Extended)'
calc[f'B{row}'] = "='All Parameters'!B25"
calc[f'C{row}'] = '=FLOOR(B14*B20/100,1)'
calc[f'C{row}'].number_format = '0'

row += 1
calc[f'A{row}'] = 'T4 (Runner)'
calc[f'B{row}'] = "='All Parameters'!B26"
calc[f'C{row}'] = '=B14-C18-C19-C20'
calc[f'C{row}'].number_format = '0'

row += 1
calc[f'A{row}'] = 'TOTAL'
calc[f'C{row}'] = '=SUM(C18:C21)'
calc[f'A{row}'].font = METRIC_FONT
calc[f'C{row}'].font = METRIC_FONT
calc[f'C{row}'].number_format = '0'

# ============ TRAILING STOP LEVELS ============
trail = wb.create_sheet('Trailing Stops', 3)
trail.column_dimensions['A'].width = 25
trail.column_dimensions['B'].width = 18
trail.column_dimensions['C'].width = 18
trail.column_dimensions['D'].width = 18

trail['A1'] = 'Trailing Stop Levels'
trail['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
trail.merge_cells('A1:D1')

trail['A3'] = 'Level'
trail['B3'] = 'Trigger (Points)'
trail['C3'] = 'Distance (Points)'
trail['D3'] = 'Net Protection'
for col in ['A', 'B', 'C', 'D']:
    trail[f'{col}3'].fill = HEADER_FILL
    trail[f'{col}3'].font = HEADER_FONT

row = 4
trail[f'A{row}'] = 'Breakeven'
trail[f'B{row}'] = "='All Parameters'!B28"
trail[f'C{row}'] = "='All Parameters'!B29/4"
trail[f'D{row}'] = '=B4-C4'
trail[f'B{row}'].number_format = '0.0'
trail[f'C{row}'].number_format = '0.00'
trail[f'D{row}'].number_format = '0.00'

row += 1
trail[f'A{row}'] = 'Trail Level 1'
trail[f'B{row}'] = "='All Parameters'!B30"
trail[f'C{row}'] = "='All Parameters'!B31"
trail[f'D{row}'] = '=B5-C5'
trail[f'B{row}'].number_format = '0.0'
trail[f'C{row}'].number_format = '0.0'
trail[f'D{row}'].number_format = '0.0'

row += 1
trail[f'A{row}'] = 'Trail Level 2'
trail[f'B{row}'] = "='All Parameters'!B32"
trail[f'C{row}'] = "='All Parameters'!B33"
trail[f'D{row}'] = '=B6-C6'
trail[f'B{row}'].number_format = '0.0'
trail[f'C{row}'].number_format = '0.0'
trail[f'D{row}'].number_format = '0.0'

row += 1
trail[f'A{row}'] = 'Trail Level 3'
trail[f'B{row}'] = "='All Parameters'!B34"
trail[f'C{row}'] = "='All Parameters'!B35"
trail[f'D{row}'] = '=B7-C7'
trail[f'B{row}'].number_format = '0.0'
trail[f'C{row}'].number_format = '0.0'
trail[f'D{row}'].number_format = '0.0'

# ============ SESSION TIMES ============
sessions = wb.create_sheet('Session Times', 4)
sessions.column_dimensions['A'].width = 25
sessions.column_dimensions['B'].width = 15
sessions.column_dimensions['C'].width = 15
sessions.column_dimensions['D'].width = 15
sessions.column_dimensions['E'].width = 30

sessions['A1'] = 'Global Trading Sessions'
sessions['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
sessions.merge_cells('A1:E1')

sessions['A3'] = 'Session'
sessions['B3'] = 'Start (Local)'
sessions['C3'] = 'End (Local)'
sessions['D3'] = 'OR Duration'
sessions['E3'] = 'Notes'
for col in ['A', 'B', 'C', 'D', 'E']:
    sessions[f'{col}3'].fill = HEADER_FILL
    sessions[f'{col}3'].font = HEADER_FONT

session_data = [
    ('NY Open (MES)', '09:30', '16:00', '5 min', 'Primary session - Eastern Time'),
    ('Australia Open (MGC)', '18:00', '01:00', '5 min', 'Sydney session - AEST'),
    ('China Open', '21:30', '04:00', '5 min', 'Shanghai session - CST'),
    ('NZ Open', '16:00', '23:00', '5 min', 'Wellington session - NZST'),
]

row = 4
for session, start, end, duration, notes in session_data:
    sessions[f'A{row}'] = session
    sessions[f'B{row}'] = start
    sessions[f'C{row}'] = end
    sessions[f'D{row}'] = duration
    sessions[f'E{row}'] = notes
    row += 1

# ============ STRATEGY COMPARISON ============
compare = wb.create_sheet('Strategy Comparison', 5)
compare.column_dimensions['A'].width = 30
compare.column_dimensions['B'].width = 18
compare.column_dimensions['C'].width = 18

compare['A1'] = 'OR vs RMA Strategy Comparison'
compare['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
compare.merge_cells('A1:C1')

compare['A3'] = 'Parameter'
compare['B3'] = 'OR Strategy'
compare['C3'] = 'RMA Strategy'
for col in ['A', 'B', 'C']:
    compare[f'{col}3'].fill = HEADER_FILL
    compare[f'{col}3'].font = HEADER_FONT

comparison_data = [
    ('Entry Type', 'Breakout (Stop Market)', 'Manual (Market Order)'),
    ('Entry Trigger', 'OR High/Low + 3 ticks', 'Shift+Click or R+Click'),
    ('Stop Calculation', 'OR Range x Multiplier', 'ATR x Multiplier'),
    ('Stop Multiplier', "='All Parameters'!B17", "='All Parameters'!B43"),
    ('T1 Calculation', 'Fixed Points', 'ATR x Multiplier'),
    ('T1 Value', "='All Parameters'!B20&\" pts\"", "='All Parameters'!B44&\"x ATR\""),
    ('T2 Calculation', 'OR Range x Multiplier', 'ATR x Multiplier'),
    ('T2 Value', "='All Parameters'!B21&\"x OR\"", "='All Parameters'!B45&\"x ATR\""),
    ('Trailing Stops', 'Yes (3 levels)', 'Yes (3 levels)'),
    ('ATR Period', 'N/A', "='All Parameters'!B42"),
]

row = 4
for param, or_val, rma_val in comparison_data:
    compare[f'A{row}'] = param
    compare[f'B{row}'] = or_val
    compare[f'C{row}'] = rma_val
    compare[f'A{row}'].font = METRIC_FONT
    compare[f'B{row}'].font = CALC_FONT
    compare[f'C{row}'].font = CALC_FONT
    row += 1

# ============ QUICK REFERENCE ============
ref = wb.create_sheet('Quick Reference', 6)
ref.column_dimensions['A'].width = 25
ref.column_dimensions['B'].width = 50

ref['A1'] = 'Quick Reference Guide'
ref['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F4E78')
ref.merge_cells('A1:B1')

ref['A3'] = 'HOTKEYS'
ref['A3'].fill = HEADER_FILL
ref['A3'].font = HEADER_FONT
ref.merge_cells('A3:B3')

hotkeys = [
    ('L', 'Enter Long OR trade'),
    ('S', 'Enter Short OR trade'),
    ('R + L/S', 'RMA Mode (hold R, then click L or S)'),
    ('Shift + Click', 'RMA entry at clicked price'),
    ('F', 'Flatten all positions'),
    ('B', 'Manual Breakeven (all positions)'),
]

row = 4
for key, action in hotkeys:
    ref[f'A{row}'] = key
    ref[f'B{row}'] = action
    ref[f'A{row}'].font = Font(name='Calibri', size=11, bold=True, color='0000FF')
    ref[f'B{row}'].font = CALC_FONT
    row += 1

row += 1
ref[f'A{row}'] = 'APEX COMPLIANCE'
ref[f'A{row}'].fill = HEADER_FILL
ref[f'A{row}'].font = HEADER_FONT
ref.merge_cells(f'A{row}:B{row}')

row += 1
compliance = [
    ('Daily Loss Limit', 'Built-in to strategy'),
    ('Max Drawdown', 'Monitored via trailing stops'),
    ('Order Rate Limiting', 'Prevents excessive orders'),
    ('Rithmic Feed', 'Optimized for fast fills'),
]

for item, status in compliance:
    ref[f'A{row}'] = item
    ref[f'B{row}'] = status
    ref[f'A{row}'].font = METRIC_FONT
    ref[f'B{row}'].font = CALC_FONT
    row += 1

wb.save('c:\\Users\\Mohammed Khalid\\OneDrive\\Desktop\\WSGTA\\Github\\universal-or-strategy\\UniversalOR_TradingDashboard_v5_13.xlsx')
print('Spreadsheet created successfully!')
